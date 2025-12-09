#!/usr/bin/env python3
"""
attach_teachers_to_timetable_final.py

Reads:
 - timetable Excel (default: timetable_tools/output_v5/All_Timetables_v5.xlsx or a provided file)
 - subjects_with_teachers.csv (contains Subject Name and teacher_id)
 - teachers.csv (contains id and name)

Produces:
 - All_Timetables_with_Teachers_fixed_v2.xlsx (timetable cells replaced with "CODE — Teacher Name (ID)")
 - missing_mappings.csv  (codes that couldn't be matched)
 - UPDATED overall_schedule.csv (adds Teacher Name + Teacher ID columns)

Usage:
    python attach_teachers_to_timetable_final.py
    python attach_teachers_to_timetable_final.py --input /path/to/All_Timetables_v5.xlsx
"""
import os
import re
import time
import csv
import argparse
from collections import defaultdict

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font

# ---------------------- Defaults ----------------------
DEFAULT_OUT_DIR = "timetable_tools/output_v5"
DEFAULT_INPUT_XLSX = os.path.join(DEFAULT_OUT_DIR, "All_Timetables_v5.xlsx")
DEFAULT_SUBJECTS_CSV = "subjects_with_teachers.csv"
FALLBACK_SUBJECTS_CSV = "subjects.csv"
DEFAULT_TEACHERS_CSV = "teachers.csv"
DEFAULT_OUTPUT_XLSX = os.path.join(DEFAULT_OUT_DIR, "All_Timetables_with_Teachers_fixed_v2.xlsx")
MISSING_LOG = "missing_mappings.csv"

WEEKDAY_TOKENS = {
    "MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN",
    "MONDAY","TUESDAY","WEDNESDAY","THURSDAY","FRIDAY","SATURDAY","SUNDAY",
    "DAY", "LUNCH"
}

# ---------------------- Utilities ----------------------
def safe_save_workbook(wb, desired_path):
    """Save workbook; if file exists and can't be overwritten, save a timestamped fallback."""
    folder = os.path.dirname(desired_path) or "."
    os.makedirs(folder, exist_ok=True)
    if os.path.exists(desired_path):
        try:
            os.remove(desired_path)
        except Exception:
            ts = time.strftime("%Y%m%d_%H%M%S")
            base, ext = os.path.splitext(desired_path)
            fallback = f"{base}_backup_{ts}{ext}"
            wb.save(fallback)
            print(f"Could not overwrite existing file. Saved to: {fallback}")
            return fallback
    wb.save(desired_path)
    return desired_path

def normalize_code(tok: str) -> str:
    s = str(tok).strip()
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"_+", "_", s)
    s = s.strip("_")
    return s.upper()

def tokenize_for_jaccard(text: str):
    if not isinstance(text, str):
        return set()
    s = re.sub(r"[^\w]", " ", text).lower()
    toks = [t for t in s.split() if len(t) > 1]
    stop = {"and","the","of","lab","laboratory","laboratories","i","ii","iii","iv","v","vi","la","1","2","3","4","5","6","7","8"}
    toks = [t for t in toks if t not in stop]
    return set(toks)

def jaccard(a:set, b:set) -> float:
    if not a or not b: return 0.0
    inter = len(a & b)
    uni = len(a | b)
    return inter / uni if uni > 0 else 0.0

# ---------------------- Read CSVs ----------------------
def read_subjects_map(subjects_csv_path):
    df = pd.read_csv(subjects_csv_path, dtype=str).fillna("")
    if "Subject Name" not in df.columns:
        for c in df.columns:
            if "subject" in c.lower():
                df = df.rename(columns={c: "Subject Name"})
                break
    subj_rows = []
    for _, r in df.iterrows():
        name = str(r.get("Subject Name","")).strip()
        tid = str(r.get("teacher_id","")).strip() if "teacher_id" in df.columns else str(r.get("teacher","")).strip() if "teacher" in df.columns else ""
        branch = str(r.get("Branch","")).strip().upper()
        if not name:
            continue
        subj_rows.append((branch, name, tid, tokenize_for_jaccard(name)))
    return subj_rows

def read_teachers(teachers_csv_path):
    df = pd.read_csv(teachers_csv_path, dtype=str).fillna("")
    cols = [c.strip() for c in df.columns]
    lower = {c.lower(): c for c in cols}
    if "id" in lower and "name" in lower:
        id_col, name_col = lower["id"], lower["name"]
    elif len(cols) >= 2:
        id_col, name_col = cols[0], cols[1]
    else:
        return {}
    mapping = {}
    for _, r in df.iterrows():
        tid = str(r.get(id_col,"")).strip()
        name = str(r.get(name_col,"")).strip()
        if tid:
            mapping[tid] = name
    return mapping

# ---------------------- Match codes to subjects ----------------------
CODE_FIND_RE = re.compile(r"\(([^\)]+)\)|\b([A-Z]{2,}[A-Z0-9_]{1,})\b")

def extract_codes_from_cell(text):
    if not isinstance(text, str):
        return []
    found = []
    for m in CODE_FIND_RE.finditer(text):
        tok = m.group(1) or m.group(2)
        if not tok: continue
        norm = normalize_code(tok)
        if norm in WEEKDAY_TOKENS: continue
        if norm not in found:
            found.append(norm)
    return found

def match_code_to_subject(code_norm, subject_rows):
    code_like = re.sub(r"_+", " ", code_norm).strip()
    code_like = re.sub(r"\b(LA|LAB|LABORATORY|LABORATOR)\b", "laboratory", code_like, flags=re.I)
    code_like = re.sub(r"\b\d+\b", "", code_like)
    code_tok = tokenize_for_jaccard(code_like)
    
    # Detect branch from code prefix (e.g. "ISE_..." -> "ISE")
    code_branch = ""
    if "_" in code_norm:
        prefix = code_norm.split("_")[0]
        if prefix in ["CSE", "ISE", "ECE", "MECH", "CIVIL", "EEE"]:
            code_branch = prefix

    best = None; best_score = 0.0; best_tid = ""
    
    for branch, subj_name, tid, subj_tokens in subject_rows:
        # Base score
        score = jaccard(code_tok, subj_tokens)
        
        # Boost score if branches match
        if code_branch and branch:
            if code_branch == branch:
                score += 0.5  # Significant boost for matching branch
            else:
                score -= 0.1  # Slight penalty for mismatching branch
        
        if score > best_score:
            best_score = score
            best = subj_name
            best_tid = tid
            
    return best, best_tid, best_score

# ---------------------- Update overall_schedule.csv ----------------------
def update_overall_schedule_csv(code_match_info):
    csv_path = os.path.join(DEFAULT_OUT_DIR, "overall_schedule.csv")
    if not os.path.exists(csv_path):
        print(f"overall_schedule.csv NOT found at: {csv_path}")
        return

    df = pd.read_csv(csv_path, dtype=str).fillna("")

    # Add new empty columns
    df["Teacher Name"] = ""
    df["Teacher ID"] = ""

    # Process each row
    for i, row in df.iterrows():
        text = row.get("Subject/Notes", "")
        codes = extract_codes_from_cell(text)

        if not codes:
            continue

        code = codes[0]
        info = code_match_info.get(code, {})

        df.at[i, "Teacher Name"] = info.get("teacher_name", "")
        df.at[i, "Teacher ID"] = info.get("teacher_id", "")

    df.to_csv(csv_path, index=False, encoding="utf-8")
    print("Updated overall_schedule.csv with Teacher Name and Teacher ID")

# ---------------------- Main processing ----------------------
def process(input_xlsx, subjects_csv, teachers_csv, output_xlsx, missing_log):
    chosen_subjects = subjects_csv if os.path.exists(subjects_csv) else (FALLBACK_SUBJECTS_CSV if os.path.exists(FALLBACK_SUBJECTS_CSV) else None)
    if not chosen_subjects:
        raise FileNotFoundError("subjects_with_teachers.csv or subjects.csv not found.")

    subject_rows = read_subjects_map(chosen_subjects)
    teachers_map = read_teachers(teachers_csv) if os.path.exists(teachers_csv) else {}

    if not os.path.exists(input_xlsx):
        raise FileNotFoundError(f"Input Excel not found: {input_xlsx}")

    wb_in = load_workbook(input_xlsx)
    all_codes = set()
    for sheet in wb_in.sheetnames:
        ws = wb_in[sheet]
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, str):
                    all_codes.update(extract_codes_from_cell(cell))

    # Build match info
    code_match_info = {}
    for code in sorted(all_codes):
        best_subj, tid, score = match_code_to_subject(code, subject_rows)
        teacher_name = teachers_map.get(tid, "") if tid else ""
        code_match_info[code] = {
            "matched_subject": best_subj or "",
            "teacher_id": tid or "",
            "teacher_name": teacher_name or "",
            "score": score
        }

    # Build Excel output
    wb_out = Workbook()
    if wb_out.active: wb_out.remove(wb_out.active)

    missing_codes = set()
    for sheet in wb_in.sheetnames:
        ws_in = wb_in[sheet]
        ws_out = wb_out.create_sheet(sheet)

        for r_idx, row in enumerate(ws_in.iter_rows(values_only=False), start=1):
            for c_idx, cell in enumerate(row, start=1):
                orig = cell.value if cell is not None else ""

                if c_idx == 1 and isinstance(orig, str) and normalize_code(orig) in WEEKDAY_TOKENS:
                    out_val = orig
                else:
                    codes = extract_codes_from_cell(orig)
                    if not codes:
                        out_val = orig
                    else:
                        lines = []
                        for code in codes:
                            info = code_match_info.get(code, {})
                            tname = info.get("teacher_name", "")
                            tid = info.get("teacher_id", "")
                            if tname:
                                lines.append(f"{code} — {tname} ({tid})")
                            elif tid:
                                lines.append(f"{code} — {tid}")
                            else:
                                lines.append(f"{code} — (no teacher)")
                                missing_codes.add(code)
                        out_val = "\n".join(lines)

                out_cell = ws_out.cell(row=r_idx, column=c_idx, value=out_val)
                out_cell.alignment = Alignment(wrap_text=True, vertical="top")

                try:
                    if cell.font and cell.font.bold:
                        out_cell.font = Font(bold=True)
                except Exception:
                    pass

        for col in ws_out.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    l = max(len(str(line)) for line in str(cell.value).split("\n"))
                    max_len = max(max_len, l)
            ws_out.column_dimensions[col_letter].width = min(max(12, max_len + 2), 50)

    saved = safe_save_workbook(wb_out, output_xlsx)
    print(f"Saved timetable with teachers to: {saved}")

    if missing_codes:
        with open(missing_log, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["missing_code", "note", "matched_subject_suggestion"])
            for code in sorted(missing_codes):
                w.writerow([code, "no teacher found", code_match_info.get(code, {}).get("matched_subject","")])
        print(f"Missing mappings written to: {missing_log}")
    else:
        print("All codes matched to teachers.")

    # 🔥 Update CSV with new Teacher Columns
    update_overall_schedule_csv(code_match_info)
    
    return saved

# ---------------------- CLI ----------------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default=DEFAULT_INPUT_XLSX)
    parser.add_argument("--subjects", default=DEFAULT_SUBJECTS_CSV)
    parser.add_argument("--teachers", default=DEFAULT_TEACHERS_CSV)
    parser.add_argument("--output", default=DEFAULT_OUTPUT_XLSX)
    parser.add_argument("--missing", default=MISSING_LOG)
    args = parser.parse_args()

    process(args.input, args.subjects, args.teachers, args.output, args.missing)
