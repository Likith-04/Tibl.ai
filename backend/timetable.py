#!/usr/bin/env python3
# auto_scheduler_final_swap.py
# Deterministic swap lab scheduler (ready-to-run).
# Ensures each batch attends every lab once per week.
# For each lab-pair (A,B) we schedule two sessions:
#  - Day1: A1->A, A2->B
#  - Day2: A1->B, A2->A
# Save next to subjects.csv (Branch,Subject Type,Subject Name[,code,credits,teacher_id])
# Run: python auto_scheduler_final_swap.py

import os
import random
import re
from collections import defaultdict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# ---------- CONFIG ----------
OUT_DIR = "timetable_tools/output_v5"
os.makedirs(OUT_DIR, exist_ok=True)

DAYS = ["MON", "TUE", "WED", "THU", "FRI"]
TIME_SLOTS = [
    "09:00-10:00", "10:00-11:00", "11:00-11:20",
    "11:20-12:20", "12:20-13:20", "13:20-14:00 (Lunch)",
    "14:00-15:00", "15:00-16:00", "16:00-17:00"
]
BLOCKED = {2, 5}                      # blocked indices (e.g., short break / lunch)
ELIGIBLE_LAB_STARTS = [0, 3, 6, 7]   # valid 2-hour lab starts (start and start+1 used)

BRANCH_SECTIONS = {
    "CSE": ["A", "B", "C"],
    "ISE": ["D", "E"],
    "ECE": ["F"]
}

LAB_POOLS = {
    "CSE": ["CSE_Lab1", "CSE_Lab2"],
    "ISE": ["ISE_Lab1", "ISE_Lab2"],
    "ECE": ["ECE_Lab1", "ECE_Lab2"]
}

PREFERRED_LAB_DAYS = ["TUE", "THU"]  # preferred lab-days; script will use these first

random.seed(42)

# ---------- HELPERS ----------
def slugify(s: str) -> str:
    s = str(s or "").strip().upper()
    s = re.sub(r"[^\w\s-]", "", s)
    s = re.sub(r"\s+", "_", s)
    return s[:40]

def load_subjects_teachers(subjects_path="subjects.csv", teachers_path="teachers.csv"):
    subs = pd.read_csv(subjects_path).fillna("")
    subs.columns = [c.strip() for c in subs.columns]

    teacher_map = {}
    if os.path.exists(teachers_path):
        teaches = pd.read_csv(teachers_path).fillna("")
        cols = list(teaches.columns)
        if any(c.lower() == "id" for c in cols) and any(c.lower() == "name" for c in cols):
            id_col = next(c for c in cols if c.lower() == "id")
            name_col = next(c for c in cols if c.lower() == "name")
            for _, r in teaches.iterrows():
                teacher_map[str(r[id_col]).strip()] = str(r[name_col]).strip()
        elif len(cols) >= 2:
            for _, r in teaches.iterrows():
                teacher_map[str(r[cols[0]]).strip()] = str(r[cols[1]]).strip()

    cols_lower = {c.lower(): c for c in subs.columns}
    def find_col(*names, default=None):
        for n in names:
            if n.lower() in cols_lower:
                return cols_lower[n.lower()]
        return default

    branch_col = find_col("Branch", "branch", default=None)
    stype_col = find_col("Subject Type", "type", "subject type", default=None)
    name_col = find_col("Subject Name", "name", "subject name", default=None)
    code_col = find_col("code", "Code", default=None)
    teacher_col = find_col("teacher_id", "teacher", "teacher id", default=None)
    credits_col = find_col("credits", default=None)

    subject_map = {}
    counters = defaultdict(int)
    for _, r in subs.iterrows():
        row = dict(r)
        branch = str(row.get(branch_col, "")).strip().upper() if branch_col else ""
        stype = str(row.get(stype_col, "")).strip().title() if stype_col else "Theory"
        name = str(row.get(name_col, "")).strip() if name_col else ""
        provided_code = str(row.get(code_col, "")).strip() if code_col else ""
        teacher_id = str(row.get(teacher_col, "")).strip() if teacher_col else ""
        credits = row.get(credits_col, "") if credits_col else ""
        if credits == "" or credits is None:
            credits = None
        else:
            try:
                credits = int(credits)
            except Exception:
                credits = None

        if provided_code:
            code = provided_code.upper()
        else:
            counters[branch] += 1
            code = f"{branch}_{slugify(name)[:20]}_{counters[branch]}"

        if credits is None:
            if stype.lower() == "theory":
                credits = 4
            elif stype.lower() == "project":
                credits = 2
            elif stype.lower() == "lab":
                credits = 0
            else:
                credits = 3

        subject_map[code] = {
            "code": code,
            "branch": branch,
            "name": name,
            "type": stype,
            "teacher_id": teacher_id,
            "credits": int(credits)
        }

    return subject_map, teacher_map

# ---------- TIMETABLE CLASS ----------
class TimeTable:
    def __init__(self, subjects, teachers):
        self.subjects = subjects
        self.teachers = teachers
        self.section_tables = {}
        self.teacher_busy = defaultdict(set)
        # batch_lab_days[(section,batch)] = set(days where batch has lab) used for constraints
        self.batch_lab_days = defaultdict(set)
        self.allocations = []

    def init_section(self, section):
        grid = {d: {i: ("", None) for i in range(len(TIME_SLOTS))} for d in DAYS}
        self.section_tables[section] = grid

    def is_free(self, section, day, slot, teacher_id=None):
        if slot in BLOCKED: return False
        if slot < 0 or slot >= len(TIME_SLOTS): return False
        cell = self.section_tables[section][day][slot]
        if cell and cell[0]: return False
        if teacher_id and (day, slot) in self.teacher_busy.get(teacher_id, set()): return False
        return True

    def mark(self, section, day, slot, subj_code, teacher_id):
        self.section_tables[section][day][slot] = (subj_code, teacher_id)
        if teacher_id:
            self.teacher_busy[teacher_id].add((day, slot))
        self.allocations.append((day, section, slot, subj_code, teacher_id))

    def assign_theory_and_project(self, section):
        branch = section.split("-")[0]
        subj_codes = [code for code, info in self.subjects.items()
                      if info["type"].lower() in ("theory", "project") and info["branch"] == branch]
        sessions = {}
        for code in subj_codes:
            credits = int(self.subjects[code].get("credits", 0) or 0)
            sessions[code] = max(1, min(credits, 6))

        per_day_load = {d: sum(1 for _ in self.section_tables[section][d].items() if _[1] and _[1][0]) for d in DAYS}
        codes = list(sessions.keys())
        random.shuffle(codes)
        subject_used_days = defaultdict(set)

        for code in codes:
            teacher = self.subjects[code].get("teacher_id", "")
            count = sessions[code]
            while count > 0:
                candidate_days = sorted(DAYS, key=lambda d: per_day_load[d])
                placed = False
                for d in candidate_days:
                    if d in subject_used_days[code]: continue
                    free_slots = [i for i in range(len(TIME_SLOTS)) if i not in BLOCKED and self.is_free(section, d, i, teacher)]
                    if not free_slots: continue
                    slot = random.choice(free_slots)
                    self.mark(section, d, slot, code, teacher)
                    per_day_load[d] += 1
                    subject_used_days[code].add(d)
                    count -= 1
                    placed = True
                    break
                if not placed:
                    placed_any = False
                    for d in DAYS:
                        free_slots = [i for i in range(len(TIME_SLOTS)) if i not in BLOCKED and self.is_free(section, d, i, teacher)]
                        if free_slots:
                            slot = random.choice(free_slots)
                            self.mark(section, d, slot, code, teacher)
                            per_day_load[d] += 1
                            subject_used_days[code].add(d)
                            count -= 1
                            placed_any = True
                            break
                    if not placed_any:
                        break

    # ---------- assign labs so each batch attends each lab once per week ----------
    def assign_labs(self, branch, section):
        try:
            df = pd.read_csv("subjects.csv", dtype=str).fillna("")
            lab_rows = df[(df.get("Branch", "").astype(str).str.strip().str.upper() == branch.strip().upper()) &
                          (df.get("Subject Type", df.get("type", "")).astype(str).str.strip().str.lower() == "lab")]
        except Exception:
            lab_subjects = [s for s in self.subjects.values()
                            if str(s.get("type", "")).strip().lower() == "lab" and s.get("branch") == branch]
        else:
            lab_subjects = []
            for _, r in lab_rows.iterrows():
                code = str(r.get("code", "")).strip()
                name = str(r.get("Subject Name", r.get("name", ""))).strip()
                tid = str(r.get("teacher_id", "")).strip()
                chosen = None
                if code and code in self.subjects:
                    chosen = self.subjects[code]
                else:
                    for c, info in self.subjects.items():
                        if info.get("branch", "").strip().upper() == branch.strip().upper() and info.get("name", "").strip().upper() == name.upper():
                            chosen = info; break
                if chosen:
                    lab_subjects.append(chosen)
                else:
                    lab_subjects.append({"code": code or f"{branch}_LAB_{len(lab_subjects)+1}",
                                         "branch": branch, "name": name, "type": "Lab", "teacher_id": tid, "credits": 1})

        if not lab_subjects:
            return

        # ensure lab rooms
        lab_pool = LAB_POOLS.get(branch, []).copy()
        if len(lab_pool) < max(2, len(lab_subjects) * 2):
            base = lab_pool[0] if lab_pool else f"{branch}_Lab"
            idx = 1
            while len(lab_pool) < max(2, len(lab_subjects) * 2):
                cand = f"{base}_{idx}"
                if cand not in lab_pool:
                    lab_pool.append(cand)
                idx += 1

        sec_letter = section.split("-")[-1]
        batch1 = f"{sec_letter}1"
        batch2 = f"{sec_letter}2"

        # candidate lab days: preferred first, then the rest
        lab_days = [d for d in PREFERRED_LAB_DAYS if d in DAYS] + [d for d in DAYS if d not in PREFERRED_LAB_DAYS]

        # clear prior lab text on lab_days
        for d in lab_days:
            for s in ELIGIBLE_LAB_STARTS:
                if s + 1 >= len(TIME_SLOTS): continue
                cell = self.section_tables[section][d][s]
                if cell and isinstance(cell, tuple) and cell[0] and "->" in str(cell[0]):
                    self.section_tables[section][d][s] = ("", None)
                    self.section_tables[section][d][s+1] = ("", None)

        # Pair up lab_subjects (two-by-two). For each pair we will schedule TWO sessions (swap) on two different days:
        # Session DayA: batch1->labA, batch2->labB
        # Session DayB: batch1->labB, batch2->labA
        pairs = []
        it = iter(lab_subjects)
        while True:
            try:
                a = next(it)
            except StopIteration:
                break
            try:
                b = next(it)
            except StopIteration:
                b = None
            pairs.append((a, b))

        # We'll try to find two different days for each pair. Use lab_days round-robin.
        day_cursor = 0
        for i, (lab_a, lab_b) in enumerate(pairs):
            # pick two distinct days for the pair, prefer days where neither batch yet has a lab
            def find_day(start_index):
                n = len(lab_days)
                for offset in range(n):
                    d = lab_days[(start_index + offset) % n]
                    if d not in self.batch_lab_days[(section, batch1)] and d not in self.batch_lab_days[(section, batch2)]:
                        return (start_index + offset) % n, d
                # fallback: return smallest-count day
                scores = []
                for idx_d, d in enumerate(lab_days):
                    cnt = (1 if d in self.batch_lab_days[(section, batch1)] else 0) + (1 if d in self.batch_lab_days[(section, batch2)] else 0)
                    scores.append((cnt, idx_d, d))
                scores.sort()
                return scores[0][1], scores[0][2]

            idx1, day1 = find_day(day_cursor)
            idx2, day2 = find_day((idx1 + 1) % len(lab_days))
            # ensure day2 != day1; if equal, pick next best
            if day2 == day1:
                for off in range(1, len(lab_days)):
                    idx_try = (idx1 + off) % len(lab_days)
                    d_try = lab_days[idx_try]
                    if d_try != day1:
                        idx2, day2 = idx_try, d_try
                        break

            day_cursor = (idx2 + 1) % len(lab_days)

            # pick rooms for lab_a and lab_b (distinct)
            used_rooms = set()
            def pick_room(name_key):
                # try substring match
                substr = "".join(re.findall(r"[A-Za-z]+", (name_key or "") ) ).lower()[:6]
                for r in lab_pool:
                    if r in used_rooms: continue
                    if substr and substr in r.lower():
                        used_rooms.add(r); return r
                for r in lab_pool:
                    if r not in used_rooms:
                        used_rooms.add(r); return r
                # fallback synthetic
                idx = 1
                while True:
                    cand = f"{branch}_Lab_{idx}"
                    if cand not in used_rooms:
                        used_rooms.add(cand); return cand
                    idx += 1

            room_a = pick_room(lab_a.get("name") if lab_a else "")
            room_b = pick_room(lab_b.get("name") if lab_b else "")

            # If lab_b is None (odd number), pick another distinct room for second session
            if lab_b is None:
                room_b = pick_room(lab_a.get("name"))

            # Two sessions: day1 and day2 with swapped mapping
            sessions = []
            # session1 mapping (day1)
            if lab_a and lab_b:
                # day1: batch1->lab_a, batch2->lab_b
                sessions.append((day1, [(batch1, room_a, lab_a.get("code")), (batch2, room_b, lab_b.get("code"))]))
                # day2: batch1->lab_b, batch2->lab_a
                sessions.append((day2, [(batch1, room_b, lab_b.get("code")), (batch2, room_a, lab_a.get("code"))]))
            else:
                # single lab present: schedule twice so both batches attend it once (in distinct rooms)
                sessions.append((day1, [(batch1, room_a, lab_a.get("code"))]))
                sessions.append((day2, [(batch2, room_b, lab_a.get("code"))]))

            # Place each session (respectful placement -> relocation -> forced overwrite)
            for day, mapping in sessions:
                placed = False
                # Try respectful placement: find start slot where all required slots are free
                for start in ELIGIBLE_LAB_STARTS:
                    if start + 1 >= len(TIME_SLOTS): continue
                    if start in BLOCKED or (start+1) in BLOCKED: continue
                    good = True
                    for (batch, room, code) in mapping:
                        # Ensure section has free slot (teacher conflicts for labs ignored here except if teacher set)
                        teacher_ids = set()
                        # if code present and teacher assigned, block teacher too
                        if code and code in self.subjects:
                            tid = self.subjects[code].get("teacher_id","")
                        else:
                            tid = None
                        if not self.is_free(section, day, start, tid) or not self.is_free(section, day, start+1, tid):
                            good = False; break
                    if not good: continue
                    # place mapping
                    for (batch, room, code) in mapping:
                        txt = f"{batch} -> {room} ({code})" if code else f"{batch} -> {room}"
                        existing = self.section_tables[section][day][start][0]
                        if existing and "->" in str(existing):
                            self.section_tables[section][day][start] = (existing + "; " + txt, None)
                        else:
                            self.section_tables[section][day][start] = (txt, None)
                        self.section_tables[section][day][start+1] = ("", None)
                        # mark teacher busy if teacher exists
                        if code and code in self.subjects:
                            tid = self.subjects[code].get("teacher_id","")
                            if tid:
                                self.teacher_busy[tid].add((day, start)); self.teacher_busy[tid].add((day, start+1))
                        # record batch day usage
                        self.batch_lab_days[(section, batch)].add(day)
                    placed = True
                    break
                if placed:
                    continue

                # Try relocation of conflicting theory classes / move blocking teachers
                moved_any = False
                for cand_start in ELIGIBLE_LAB_STARTS:
                    if cand_start + 1 >= len(TIME_SLOTS): continue
                    if cand_start in BLOCKED or (cand_start+1) in BLOCKED: continue
                    target_days = [day] + [d for d in DAYS if d != day]
                    success = False
                    for cand_day in target_days:
                        can_place = True
                        for (batch, room, code) in mapping:
                            # teacher check if code maps to teacher
                            tid = self.subjects.get(code,{}).get("teacher_id","") if code else None
                            if not self.is_free(section, cand_day, cand_start, tid) or not self.is_free(section, cand_day, cand_start+1, tid):
                                can_place = False; break
                        if not can_place: continue
                        # place mappings
                        for (batch, room, code) in mapping:
                            txt = f"{batch} -> {room} ({code})" if code else f"{batch} -> {room}"
                            existing = self.section_tables[section][cand_day][cand_start][0]
                            if existing and "->" in str(existing):
                                self.section_tables[section][cand_day][cand_start] = (existing + "; " + txt, None)
                            else:
                                self.section_tables[section][cand_day][cand_start] = (txt, None)
                            self.section_tables[section][cand_day][cand_start+1] = ("", None)
                            if code and code in self.subjects:
                                tid = self.subjects[code].get("teacher_id","")
                                if tid:
                                    self.teacher_busy[tid].add((cand_day, cand_start)); self.teacher_busy[tid].add((cand_day, cand_start+1))
                            self.batch_lab_days[(section, batch)].add(cand_day)
                        success = True
                        break
                    if success:
                        moved_any = True
                        break

                    # attempt to move blocking teachers' classes around (same logic as previous versions)
                    blocking = set()
                    for tid, slots in list(self.teacher_busy.items()):
                        if (day, cand_start) in slots or (day, cand_start+1) in slots:
                            blocking.add(tid)
                    moved_flag = False
                    for blocking_tid in blocking:
                        conflict_slots = [(d,s) for (d,s) in list(self.teacher_busy.get(blocking_tid, set()))
                                          if d == day and (s == cand_start or s == cand_start+1)]
                        for conf_day, conf_slot in conflict_slots:
                            for sec2, grid2 in self.section_tables.items():
                                cell = grid2[conf_day][conf_slot]
                                if cell and isinstance(cell, tuple) and cell[0] and cell[1] == blocking_tid:
                                    subj_to_move = cell[0]
                                    for alt_day in DAYS:
                                        for alt_slot in range(len(TIME_SLOTS)):
                                            if alt_slot in BLOCKED: continue
                                            if self.is_free(sec2, alt_day, alt_slot, blocking_tid):
                                                self.section_tables[sec2][alt_day][alt_slot] = (subj_to_move, blocking_tid)
                                                self.section_tables[sec2][conf_day][conf_slot] = ("", None)
                                                self.teacher_busy[blocking_tid].discard((conf_day, conf_slot))
                                                self.teacher_busy[blocking_tid].add((alt_day, alt_slot))
                                                moved_flag = True
                                                break
                                        if moved_flag: break
                                if moved_flag: break
                            if moved_flag:
                                # after move attempt place mappings
                                if all(self.is_free(section, day, cand_start+offs, "") for offs in (0,1)):
                                    for (batch, room, code) in mapping:
                                        txt = f"{batch} -> {room} ({code})" if code else f"{batch} -> {room}"
                                        existing = self.section_tables[section][day][cand_start][0]
                                        if existing and "->" in str(existing):
                                            self.section_tables[section][day][cand_start] = (existing + "; " + txt, None)
                                        else:
                                            self.section_tables[section][day][cand_start] = (txt, None)
                                        self.section_tables[section][day][cand_start+1] = ("", None)
                                        if code and code in self.subjects:
                                            tid = self.subjects[code].get("teacher_id","")
                                            if tid:
                                                self.teacher_busy[tid].add((day, cand_start)); self.teacher_busy[tid].add((day, cand_start+1))
                                        self.batch_lab_days[(section, batch)].add(day)
                                    moved_any = True
                                    break
                            if moved_any: break
                        if moved_any: break
                    if moved_any:
                        break

                if moved_any:
                    placed = True
                    continue

                # 3) FORCE OVERWRITE (last resort) - try to move blocking teacher then overwrite
                forced_done = False
                for fstart in ELIGIBLE_LAB_STARTS:
                    if fstart + 1 >= len(TIME_SLOTS): continue
                    if fstart in BLOCKED or (fstart+1) in BLOCKED: continue
                    cell = self.section_tables[section][day][fstart]
                    if cell and isinstance(cell, tuple) and cell[1]:
                        blocking_tid = cell[1]
                        moved = False
                        for alt_day in DAYS:
                            for alt_slot in range(len(TIME_SLOTS)):
                                if alt_slot in BLOCKED: continue
                                if self.is_free(section, alt_day, alt_slot, blocking_tid):
                                    subj_to_move = cell[0]
                                    self.section_tables[section][alt_day][alt_slot] = (subj_to_move, blocking_tid)
                                    self.section_tables[section][day][fstart] = ("", None)
                                    self.teacher_busy[blocking_tid].discard((day, fstart))
                                    self.teacher_busy[blocking_tid].add((alt_day, alt_slot))
                                    moved = True
                                    break
                            if moved: break
                    # write mapping overwriting any theory entries
                    for (batch, room, code) in mapping:
                        txt = f"{batch} -> {room} ({code})" if code else f"{batch} -> {room}"
                        existing = self.section_tables[section][day][fstart][0]
                        if existing and "->" in str(existing):
                            self.section_tables[section][day][fstart] = (existing + "; " + txt, None)
                        else:
                            self.section_tables[section][day][fstart] = (txt, None)
                        self.section_tables[section][day][fstart+1] = ("", None)
                        self.batch_lab_days[(section, batch)].add(day)
                    forced_done = True
                    break
                if forced_done:
                    placed = True
                    continue

                # 4) last-resort unassigned markers (shouldn't happen)
                for uday in DAYS:
                    for ustart in ELIGIBLE_LAB_STARTS:
                        if ustart + 1 >= len(TIME_SLOTS): continue
                        if ustart in BLOCKED or (ustart+1) in BLOCKED: continue
                        if not self.section_tables[section][uday][ustart][0] and not self.section_tables[section][uday][ustart+1][0]:
                            for (batch, room, code) in mapping:
                                self.section_tables[section][uday][ustart] = (f"**UNASSIGNED-LAB {code}({batch})**", None)
                                self.section_tables[section][uday][uday] = ("", None)
                                self.batch_lab_days[(section, batch)].add(uday)
                            placed = True
                            break
                    if placed: break
                if placed:
                    continue

        # end pairs loop

    def export_csvs(self):
        overall = []
        subj_counts = defaultdict(int)

        for sec, grid in self.section_tables.items():
            branch = sec.split("-")[0]
            sec_letter = sec.split("-")[-1]
            for day in DAYS:
                for idx in range(len(TIME_SLOTS)):
                    if idx in BLOCKED: continue
                    cell = grid[day].get(idx, ("", None))
                    if not cell or not cell[0]: continue
                    val = cell[0] if isinstance(cell, tuple) else cell

                    if isinstance(val, str) and "->" in val:
                        parts = [p.strip() for p in val.split(";")]
                        for part in parts:
                            left, right = (part.split("->", 1) + [""])[:2]
                            batch_name = left.strip()
                            room_and_sub = right.strip()
                            subj_code = None
                            if "(" in room_and_sub and ")" in room_and_sub:
                                subj_code = room_and_sub.split("(")[-1].split(")")[0].strip()
                            overall.append({
                                "Day": day, "Branch": branch, "Section": sec_letter,
                                "Batch": batch_name, "Time": TIME_SLOTS[idx],
                                "Activity": "Lab", "Room": room_and_sub.split("(")[0].strip(),
                                "Subject/Notes": subj_code if subj_code else room_and_sub
                            })
                            if subj_code:
                                subj_counts[subj_code] += 2
                    else:
                        if isinstance(cell, tuple):
                            subj_code, tid = cell
                        else:
                            subj_code = str(val).strip(); tid = None
                        tname = self.teachers.get(tid, "") if tid else ""
                        overall.append({
                            "Day": day, "Branch": branch, "Section": sec_letter,
                            "Batch": f"{sec_letter}1 & {sec_letter}2",
                            "Time": TIME_SLOTS[idx],
                            "Activity": "Theory/Project",
                            "Room": f"{sec_letter}-Classroom",
                            "Subject/Notes": subj_code
                        })
                        if subj_code:
                            subj_counts[subj_code] += 1

        df_overall = pd.DataFrame(overall)
        if not df_overall.empty:
            df_overall = df_overall.sort_values(by=["Branch","Section","Day","Time","Batch"])
        df_overall.to_csv(os.path.join(OUT_DIR, "overall_schedule.csv"), index=False, encoding="utf-8")
        print(f"✅ Wrote overall CSV: {os.path.join(OUT_DIR, 'overall_schedule.csv')}")

        alloc_rows = [{"Subject": k, "TotalPeriods": v} for k, v in subj_counts.items()]
        pd.DataFrame(alloc_rows).to_csv(os.path.join(OUT_DIR, "allocation_summary.csv"), index=False)
        print(f"✅ Wrote allocation summary: {os.path.join(OUT_DIR, 'allocation_summary.csv')}")

        wb = Workbook()
        if wb.active: wb.remove(wb.active)
        for sec, grid in self.section_tables.items():
            ws = wb.create_sheet(sec)
            header = ["Day"] + TIME_SLOTS
            ws.append(header)
            for col in range(1, len(header) + 1):
                c = ws.cell(row=1, column=col)
                c.font = Font(bold=True)
            for row_idx, day in enumerate(DAYS, start=2):
                row = [day]
                for idx in range(len(TIME_SLOTS)):
                    if idx in BLOCKED:
                        row.append(""); continue
                    cell = grid[day].get(idx, ("", None))
                    if not cell or not cell[0]:
                        row.append(""); continue
                    val = cell[0] if isinstance(cell, tuple) else cell
                    if isinstance(val, str) and "->" in val:
                        parts = [p.strip() for p in val.split(";")]
                        pretty = []
                        for p in parts:
                            left, right = (p.split("->",1) + [""])[:2]
                            pretty.append(f"{left.strip()} -> {right.strip()}")
                        row.append("\n".join(pretty))
                    else:
                        if isinstance(cell, tuple):
                            subj_code, tid = cell
                            tname = self.teachers.get(tid, "") if tid else ""
                            txt = subj_code
                        else:
                            txt = str(val)
                        row.append(txt)
                ws.append(row)
            for col_idx in range(1, len(header) + 1):
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                ws.column_dimensions[col_letter].width = 20 if col_idx > 1 else 12
                for r in range(1, len(DAYS) + 2):
                    c = ws.cell(row=r, column=col_idx)
                    c.alignment = Alignment(wrap_text=True, vertical="top")
        excel_path = os.path.join(OUT_DIR, "All_Timetables_v5.xlsx")
        wb.save(excel_path)
        print(f"✅ Wrote Excel timetables: {excel_path}")

# ---------- MAIN ----------
def main():
    # Prefer subjects_with_teachers.csv if it exists, as it contains teacher constraints
    if os.path.exists("subjects_with_teachers.csv"):
        subj_csv = "subjects_with_teachers.csv"
        print(f"[INFO] Using {subj_csv} for scheduling (includes teacher constraints).")
    else:
        subj_csv = "subjects.csv"
        print(f"[INFO] Using {subj_csv} for scheduling (no teacher constraints).")

    teacher_csv = "teachers.csv"

    if not os.path.exists(subj_csv):
        raise FileNotFoundError(f"Required file not found: {subj_csv}")

    subject_map, teacher_map = load_subjects_teachers(subj_csv, teacher_csv)
    tt = TimeTable(subject_map, teacher_map)

    for branch, secs in BRANCH_SECTIONS.items():
        for s in secs:
            sec_code = f"{branch}-{s}"
            tt.init_section(sec_code)

    for branch, secs in BRANCH_SECTIONS.items():
        for s in secs:
            sec_code = f"{branch}-{s}"
            print(f"[SCHEDULE] Processing {sec_code}")
            tt.assign_theory_and_project(sec_code)
            tt.assign_labs(branch, sec_code)

    tt.export_csvs()
    print("All scheduling complete.")

if __name__ == "__main__":
    main()
